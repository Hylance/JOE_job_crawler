from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import urllib2
import urlparse
from bs4 import NavigableString, BeautifulSoup, Comment
import re
import requests
import json

def find_current_string(current, next, next_level):
    if next is next_level or next_level is None:
        return current
    if isinstance(next.string, NavigableString) and "\n" not in next \
            and not isinstance(next, Comment):
        if current:
            current = current + "/" + next
        else:
            current = next
        if u'/' in current[-2]:
            current = current[:-2]
    return find_current_string(current, next.next_sibling, next_level)


def short_url(long_url):
    headers = {"Content-type": "application/json"}
    URL = "https://www.googleapis.com/urlshortener/v1/url?key=AIzaSyCxyUqgDd6I-j6ZJEv3Q5YDvvbyrpaI7F0"
    r = requests.post(URL, headers=headers, data=str({'longUrl': str(long_url)}))
    json_data = json.loads(r.text)
    try:
        return json_data['id']
    except:
        return long_url

def add_job_listing(html, jobLists):
    """
    :type jobLists: list
    """
    soup1 = BeautifulSoup(html, 'html.parser', from_encoding='utf-8')
    headings = soup1.find_all('h6', class_='listing-item-header-title')
    page_url = "https://www.aeaweb.org"
    for heading in headings:
        new_url = heading.find('a')['href']
        new_full_url = urlparse.urljoin(page_url, new_url)
        jobLists.append(new_full_url)
    return jobLists


def parse_job(job, joe_ids):

    new_data = []
    soup = BeautifulSoup(job, 'html.parser', from_encoding='utf-8')
    employer = soup.find('h3', class_='title').get_text()
    job_id = soup.find('div', text=re.compile('^JOE ID Number:')).get_text()
    if job_id in joe_ids:
        return new_data
    dead_line = soup.find('div', style='margin-bottom:10px; font-style: italic; padding-left: 15px;')
    if dead_line:
        dead_line = dead_line.get_text()[22:]
    new_data.append(employer)
    new_data.append(dead_line)
    new_data.append(job_id)
    dialog_text = soup.find('div', class_='dialog_text')
    texts = dialog_text.find_all('span', class_='short-desc-title', limit=5)
    for i in range(0, 4):
        current = ""
        next_element = texts[i].next_sibling
        if len(texts) is 4 and i is 3:
            terminal = None
        else:
            terminal = texts[i + 1]
        current = find_current_string(current, next_element, terminal)
        new_data.append(current)
    return new_data


if __name__ == "__main__":
    try:
        wb = load_workbook("JOE_Job_List.xlsx")
    except:
        wb = Workbook()
        ws = wb.active
        firstColumn = ["Employer", "Deadline", "JOE ID", "Job Title", "Section", "Location", "JEL Classifications",
                       "Apply Link", "Applied", "Interviewed", "Fly Out", "Received Offer", "Fuckee"]
        for i in range(1, 1 + len(firstColumn)):
            ws.cell(row=1, column=i).value = firstColumn[i - 1]
    rootURL = ["https://goo.gl/BQ3ZdX"]
    '''https://goo.gl/GoHxac'''
    "https://goo.gl/BQ3ZdX"
    ws = wb.active
    joeID = ws['C']
    jobLists = []
    ws1 = None
    old_joe_ids = []
    for joe_id in joeID:
        old_joe_ids.append(str(joe_id.value))
    if ws["A2"].value:
        ws1 = wb.create_sheet(str(datetime.date.today()))
        ws = ws1
    for url in rootURL:
        html = urllib2.urlopen(url).read()
        jobLists = add_job_listing(html, jobLists)
    for i in range(0, len(jobLists)):
        new_data = []
        start = 'A' + str(i + 2)
        end = 'H' + str(i + 2)
        new_data = parse_job(urllib2.urlopen(jobLists[i]).read(), old_joe_ids)
        if not new_data:
            break
        new_data.append(short_url(jobLists[i]))
        cell_range = ws[start : end]
        for j in range(0, len(cell_range[0])):
            cell_range[0][j].value = new_data[j]
    wb.save("JOE_Job_List.xlsx")
